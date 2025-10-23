import shutil
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

print_output = False
def _print(output):
    if print_output:
        print(output)

def WriteDataToExcel(data, output_file_name, template_location = None):
    ## TODO: update this function to use all_nodes, arcs only. all the data necessary should be in those.
    if template_location is None:
        template_location = f'{os.path.dirname(os.path.abspath(__file__))}/assets/pareto_input_template.xlsx'

    # some defaults:
    treatment_technologies = {
        "CB": {"TreatmentOperationalCost": 0.2, "TreatmentEfficiency": 0.95, "RemovalEfficiency": 0, "DesalinationTechnologies": 0, "TreatmentExpansionCost": 75},
        "CB-EV": {"TreatmentOperationalCost": 0.3, "TreatmentEfficiency": 0.95, "RemovalEfficiency": 0, "DesalinationTechnologies": 0, "TreatmentExpansionCost": 100},
        "MVC": {"TreatmentOperationalCost": 0.5, "TreatmentEfficiency": 0.5, "RemovalEfficiency": 0.99, "DesalinationTechnologies": 1, "TreatmentExpansionCost": 1000},
        "MD": {"TreatmentOperationalCost": 1.0, "TreatmentEfficiency": 0.5, "RemovalEfficiency": 0.99, "DesalinationTechnologies": 1, "TreatmentExpansionCost": 500},
    }

    treatment_capacities = {
        "J0": {"TreatmentExpansionLeadTime": 0, "TreatmentCapacityIncrements": 0},
        "J1": {"TreatmentExpansionLeadTime": 68, "TreatmentCapacityIncrements": 10000},
        "J2": {"TreatmentExpansionLeadTime": 70, "TreatmentCapacityIncrements": 20000},
        "J3": {"TreatmentExpansionLeadTime": 72, "TreatmentCapacityIncrements": 50000},
    }

    pipeline_diameters = {
        "D0": {"PipelineCapacityIncrements": 0, "PipelineDiameterValues": 0, "PipelineExpansionLeadTime_Capac": 0},
        "D4": {"PipelineCapacityIncrements": 14286, "PipelineDiameterValues": 4, "PipelineExpansionLeadTime_Capac": 1},
        "D6": {"PipelineCapacityIncrements": 35714, "PipelineDiameterValues": 6, "PipelineExpansionLeadTime_Capac": 2},
        "D8": {"PipelineCapacityIncrements": 42857, "PipelineDiameterValues": 8, "PipelineExpansionLeadTime_Capac": None},
        "D12": {"PipelineCapacityIncrements": 50000, "PipelineDiameterValues": 12, "PipelineExpansionLeadTime_Capac": None},
    }

    storage_capacities = {
        "C0": {"StorageCapacityIncrements": 0, "StorageExpansionLeadTime": 0},
        "C1": {"StorageCapacityIncrements": 50000, "StorageExpansionLeadTime": 88},
        "C2": {"StorageCapacityIncrements": 100000, "StorageExpansionLeadTime": 89},
        "C3": {"StorageCapacityIncrements": 350000, "StorageExpansionLeadTime": 90},
    }

    injection_capacities = {
        "I0": {"DisposalCapacityIncrements": 0, "DisposalExpansionLeadTime": 0},
        "I1": {"DisposalCapacityIncrements": 7143, "DisposalExpansionLeadTime": 45},
        "I2": {"DisposalCapacityIncrements": 14286, "DisposalExpansionLeadTime": None},
        "I3": {"DisposalCapacityIncrements": 50000, "DisposalExpansionLeadTime": None},
    }

    defaults = {
        "TreatmentTechnologies": treatment_technologies,
        "TreatmentCapacities": treatment_capacities,
        "PipelineDiameters": pipeline_diameters,
        "StorageCapacities": storage_capacities,
        "InjectionCapacities": injection_capacities,
    }


    # input_path = "./assets/pareto_input_template.xlsx"
    excel_path = f'{output_file_name}.xlsx'
    _print(f'writing data to excel at {excel_path}')

    ## step 1: copy pareto_input_template to new file for writing
    shutil.copyfile(template_location, excel_path)

    ## step 2: open excel workbook
    wb = load_workbook(excel_path, data_only=True)

    ## step 3: add nodes
    node_keys = [
        'ProductionPads', 'CompletionsPads', 'SWDSites', 'FreshwaterSources', 
        'StorageSites', 'TreatmentSites', 'NetworkNodes', "ReuseOptions"
    ]

    column = 1
    for node_key in node_keys:
        row = 2
        ws = wb[node_key]
        for node in data[node_key]:
            _print(f'{node_key}: adding {node}')
            cellLocation = f'{get_column_letter(column)}{row}'
            ws[cellLocation] = node
            row+=1

    ## step 4: add arcs
    piped_arcs = {
        "PNA": ["ProductionPads", "NetworkNodes"],
        "CNA": ["CompletionsPads", "NetworkNodes"],
        "CCA": ["CompletionsPads", "CompletionsPads"],
        "NNA": ["NetworkNodes", "NetworkNodes"],
        "NCA": ["NetworkNodes", "CompletionsPads"],
        "NKA": ["NetworkNodes", "SWDSites"],
        "NRA": ["NetworkNodes", "TreatmentSites"],
        "NSA": ["NetworkNodes", "StorageSites"],
        "NOA": ["NetworkNodes", "ReuseOptions"],
        "SNA": ["StorageSites", "NetworkNodes"],
        "SOA": ["StorageSites", "ReuseOptions"],
        "FCA": ["FreshwaterSources", "CompletionsPads"],
        "RCA": ["TreatmentSites", "CompletionsPads"],
        "RSA": ["TreatmentSites", "StorageSites"],
        "SCA": ["StorageSites", "CompletionsPads"],
        "RNA": ["TreatmentSites", "NetworkNodes"],
        "ROA": ["TreatmentSites", "ReuseOptions"],
    }

    for piped_arc in piped_arcs: ## assume all connections on map are for pipes
        ws = wb[piped_arc]
        piped_arc_node1 = piped_arcs[piped_arc][0]
        piped_arc_node2 = piped_arcs[piped_arc][1]
        column = 1
        row = 3
        row_nodes = []
        if len(data[piped_arc_node1]) > 0 and len(data[piped_arc_node2]) > 0:
            _print(f'{piped_arc}: adding {piped_arc_node1}')
            for node in data[piped_arc_node1]:
                cellLocation = f'{get_column_letter(column)}{row}'
                ws[cellLocation] = node
                row_nodes.append(node)
                row+=1
            column = 2
            row = 2
            _print(f'{piped_arc}: adding {piped_arc_node2}')
            for node in data[piped_arc_node2]:
                cellLocation = f'{get_column_letter(column)}{row}'
                ws[cellLocation] = node
                # _print('checking for connections')
                ind = 3
                for row_node in row_nodes:
                    if row_node in data["connections"]["all_connections"]:
                        if node in data["connections"]["all_connections"][row_node]:
                            # _print(f'adding connection for {row_node}:{node}')
                            cellLocation = f'{get_column_letter(column)}{ind}'
                            # _print(f'adding to cell location: {cellLocation}')
                            ws[cellLocation] = 1

                    ind+=1
                column+=1
        else:
            _print(f'removing header for {piped_arc}')
            cellLocation = f'{get_column_letter(1)}{2}'
            ws[cellLocation] = None

    trucked_arcs = {
        "PCT": ["ProductionPads", "CompletionsPads"],
        "FCT": ["FreshwaterSources", "CompletionsPads"],
        "PKT": ["ProductionPads", "SWDSites"],
        "CKT": ["CompletionsPads", "SWDSites"],
        "CCT": ["CompletionsPads", "CompletionsPads"],
        "CST": ["CompletionsPads", "StorageSites"],
        "RST": ["TreatmentSites", "StorageSites"],
        "ROT": ["TreatmentSites", "ReuseOptions"],
        "SOT": ["StorageSites", "ReuseOptions"],
    }

    for trucked_arc in trucked_arcs:
        ws = wb[trucked_arc]
        trucked_arc_node1 = trucked_arcs[trucked_arc][0]
        trucked_arc_node2 = trucked_arcs[trucked_arc][1]
        column = 1
        row = 3
        if len(data[trucked_arc_node1]) > 0 and len(data[trucked_arc_node2]) > 0:
            _print(f'{trucked_arc}: adding {trucked_arc_node1}')
            for node in data[trucked_arc_node1]:
                cellLocation = f'{get_column_letter(column)}{row}'
                ws[cellLocation] = node
                row+=1
            column = 2
            row = 2
            _print(f'{trucked_arc}: adding {trucked_arc_node2}')
            for node in data[trucked_arc_node2]:
                cellLocation = f'{get_column_letter(column)}{row}'
                ws[cellLocation] = node
                ind = 3
                column+=1
        else:
            _print(f'removing header for {trucked_arc}')
            cellLocation = f'{get_column_letter(1)}{2}'
            ws[cellLocation] = None

    ## step 5: add elevations:
    elevation_nodes = [
        'ProductionPads', 'CompletionsPads', 'SWDSites', 'FreshwaterSources', 
        'StorageSites', 'TreatmentSites', 'ReuseOptions', 'NetworkNodes'
    ]

    column = 1
    row = 3
    ws = wb["Elevation"]
    for node_key in elevation_nodes:
        for node in data[node_key]:
            _print(f'{node_key}: adding {node} elevation')
            nodeCellLocation = f'{get_column_letter(column)}{row}'
            valueCellLocation = f'{get_column_letter(column+1)}{row}'
            ws[nodeCellLocation] = node
            current_node = data[node_key][node]
            altitude = current_node.get("altitude", None)
            try:
                ws[valueCellLocation] = float(altitude)
            except:
                _print(f'unable to convert elevation to float. adding it as is: {altitude}')
                ws[valueCellLocation] = altitude
            row+=1

    ## step 6: add forecasts (with empty values)
    forecast_tabs = {
        "CompletionsDemand": ["CompletionsPads"],
        "PadRates": ["ProductionPads"],
        "FlowbackRates": ["CompletionsPads"],
        "WellPressure": ["ProductionPads", "CompletionsPads"],
        "ReuseMinimum": ["ReuseOptions"],
        "ReuseCapacity": ["ReuseOptions"],
        "FreshwaterSourcingAvailability": ["FreshwaterSources"],
        "DisposalOperatingCapacity": ["SWDSites"], ## AUTOFILL with ones?
    }

    column = 1
    for forecast_tab_key in forecast_tabs:
        ws = wb[forecast_tab_key]
        row = 3
        for node_key in forecast_tabs[forecast_tab_key]:
            _print(f'{forecast_tab_key}: adding {node_key}')
            for node in data[node_key]:
                cellLocation = f'{get_column_letter(column)}{row}'
                ws[cellLocation] = node
                row+=1

    ## step 7: add initial pipelines, capacities, ...
    initial_pipeline_tabs = {
        "InitialPipelineCapacity": [
            ["ProductionPads", "CompletionsPads", "NetworkNodes", "StorageSites", "FreshwaterSources", "TreatmentSites"], 
            ["NetworkNodes", "SWDSites", "TreatmentSites", "StorageSites", "ReuseOptions", "CompletionsPads"],
        ],
        "InitialPipelineDiameters": [
            ["ProductionPads", "CompletionsPads", "NetworkNodes", "StorageSites", "FreshwaterSources", "TreatmentSites"], 
            ["NetworkNodes", "SWDSites", "TreatmentSites", "StorageSites", "ReuseOptions", "CompletionsPads"],
        ],
        "PipelineOperationalCost": [
            ["ProductionPads", "CompletionsPads", "NetworkNodes", "StorageSites", "FreshwaterSources", "TreatmentSites"], 
            ["NetworkNodes", "SWDSites", "TreatmentSites", "StorageSites", "ReuseOptions", "CompletionsPads"],
        ],
        "PipelineExpansionDistance": [
            ["ProductionPads", "CompletionsPads", "NetworkNodes", "StorageSites", "FreshwaterSources", "TreatmentSites"], 
            ["NetworkNodes", "SWDSites", "TreatmentSites", "StorageSites", "ReuseOptions", "CompletionsPads"],
        ],
        "TruckingTime": [
            ["ProductionPads", "CompletionsPads"], 
            ["SWDSites"],
        ],
    }
    
    for initial_pipeline_tab in initial_pipeline_tabs:
        ws = wb[initial_pipeline_tab]

        # write row indexes
        column = 1
        row = 3
        node_keys = initial_pipeline_tabs[initial_pipeline_tab][0]
        for node_key in node_keys:
            _print(f'{initial_pipeline_tab}: adding {node_key}')
            for node in data[node_key]:
                cellLocation = f'{get_column_letter(column)}{row}'
                ws[cellLocation] = node
                row+=1
        
        # write column indexes
        node_keys = initial_pipeline_tabs[initial_pipeline_tab][1]
        column = 2
        row = 2
        for node_key in node_keys:
            _print(f'{initial_pipeline_tab}: adding {node_key}')
            for node in data[node_key]:
                cellLocation = f'{get_column_letter(column)}{row}'
                ws[cellLocation] = node
                column+=1

    initial_capacity_tabs = {
        "InitialDisposalCapacity": "SWDSites",
        "InitialStorageCapacity": "StorageSites",
        "CompletionsPadStorage": "CompletionsPads",
        "PadOffloadingCapacity": "CompletionsPads",
        "NodeCapacities": "NetworkNodes",
    }
    column = 1
    for initial_capacity_tab in initial_capacity_tabs:
        ws = wb[initial_capacity_tab]
        node_key = initial_capacity_tabs[initial_capacity_tab]
        row = 3
        _print(f'{initial_capacity_tab}: adding {node_key}')
        for node in data[node_key]:
            cellLocation = f'{get_column_letter(column)}{row}'
            ws[cellLocation] = node
            row+=1

    single_value_tabs = {
        "DisposalOperationalCost": ["SWDSites"], ## AUTOFILL 0.35?
        "ReuseOperationalCost": ["CompletionsPads"], ## AUTOFILL 0?
        "FreshSourcingCost": ["FreshwaterSources"],
        "TruckingHourlyCost": ["ProductionPads", "CompletionsPads", "FreshwaterSources"],
        "DesalinationSites": ["TreatmentSites"], ## AUTOFILL 0's or 1's?
        "BeneficialReuseCredit": ["ReuseOptions"],
        "CompletionsPadOutsideSystem": ["CompletionsPads"],

    }

    column = 1
    for single_value_tab in single_value_tabs:
        ws = wb[single_value_tab]
        row = 3
        for node_key in single_value_tabs[single_value_tab]:
            _print(f'{single_value_tab}: adding {node_key}')
            for node in data[node_key]:
                cellLocation = f'{get_column_letter(column)}{row}'
                ws[cellLocation] = node
                row+=1

    water_quality_tabs = { ## AUTOFILL all these to ~150,000.00?
        "PadWaterQuality": ["ProductionPads", "CompletionsPads"],
        "StorageInitialWaterQuality": ["StorageSites"],
        "PadStorageInitialWaterQuality": ["CompletionsPads"],
    }

    column = 1
    for water_quality_tab in water_quality_tabs:
        ws = wb[water_quality_tab]
        row = 3
        for node_key in water_quality_tabs[water_quality_tab]:
            _print(f'{water_quality_tab}: adding {node_key}')
            for node in data[node_key]:
                cellLocation = f'{get_column_letter(column)}{row}'
                ws[cellLocation] = node
                row+=1


    ## step 8: add tabs that rely on TreatmentTechnologies, Capacities, Diameters
    expansion_tabs = {
        "DisposalExpansionCost": {"node": "SWDSites", "default": "InjectionCapacities"},
        "StorageExpansionCost": {"node": "StorageSites", "default": "StorageCapacities"},
        "DisposalExpansionLeadTime": {"node": "SWDSites", "default": "InjectionCapacities"},
        "StorageExpansionLeadTime": {"node": "StorageSites", "default": "StorageCapacities"},
    }

    for expansion_tab in expansion_tabs:
        ws = wb[expansion_tab]
        node_key = expansion_tabs[expansion_tab]["node"]
        # add column headers
        column = 2
        for header in defaults[expansion_tabs[expansion_tab]["default"]]:
            columnHeaderCellLocation = f'{get_column_letter(column)}{2}'
            ws[columnHeaderCellLocation] = header
            column+=1
        row = 3
        _print(f'{expansion_tab}: adding {node_key}')
        for node in data[node_key]:
            cellLocation = f'{get_column_letter(1)}{row}'
            ws[cellLocation] = node
            row+=1
    
    capacity_increments_tabs = {
        "DisposalCapacityIncrements": {"default": "InjectionCapacities"},
        "StorageCapacityIncrements": {"default": "StorageCapacities"},
        "PipelineDiameterValues": {"default": "PipelineDiameters"},
        "PipelineCapacityIncrements": {"default": "PipelineDiameters"},
        "DesalinationTechnologies": {"default": "TreatmentTechnologies"},
    }

    for capacity_increments_tab in capacity_increments_tabs:
        ws = wb[capacity_increments_tab]
        default_values = defaults[capacity_increments_tabs[capacity_increments_tab]["default"]]
        row = 3
        try:
            for each in default_values:
                _print(f'{capacity_increments_tab}: adding {each}')
                keyCellLocation = f'{get_column_letter(1)}{row}'
                valueCellLocation = f'{get_column_letter(2)}{row}'
                ws[keyCellLocation] = each
                ws[valueCellLocation] = default_values[each][capacity_increments_tab]
                row+=1
        except Exception as e:
            _print(f'unable to add {capacity_increments_tab}: {e}')

    capacity_increments_tabs = {
        "TreatmentCapacityIncrements": {"node": "SWDSites", "default": "TreatmentCapacities"},
    }

    for capacity_increments_tab in capacity_increments_tabs:
        ws = wb[capacity_increments_tab]
        node_key = capacity_increments_tabs[capacity_increments_tab]["node"]
        row = 3
        for tech in treatment_technologies:
            techCellLocation = f'{get_column_letter(1)}{row}'
            ws[techCellLocation] = tech
            default_values = defaults[capacity_increments_tabs[capacity_increments_tab]["default"]]
            column = 2
            for each in default_values:
                _print(f'{capacity_increments_tab}: adding {each}')

                # add header
                headerCellLocation = f'{get_column_letter(column)}{2}'
                ws[headerCellLocation] = each

                valueCellLocation = f'{get_column_letter(column)}{row}'
                ws[valueCellLocation] = default_values[each][capacity_increments_tab]
                column+=1
            row+=1

    ## CODE TO GET TREATMENT TECHNOLOGIES. WE ARE DEFAULTING TO CB, CB-EV, MVC, and MD; so hard code those with corresponding values for now

    # treatment_technologies = []
    # ws = wb["TreatmentTechnologies"]
    # column = 1
    # row = 2
    # cellLocation = f'{get_column_letter(column)}{row}'
    # technology = ws[cellLocation].value
    # while technology is not None and technology != "" and row < 100: # just wanna make sure this doesnt get caught in an infinite loop
    #     _print(f'adding technology: {technology}')
    #     treatment_technologies.append(technology)
    #     row += 1
    #     cellLocation = f'{get_column_letter(column)}{row}'
    #     technology = ws[cellLocation].value

    tabs = {
        "InitialTreatmentCapacity": "TreatmentSites"
    }
    
    for tab in tabs:
        ws = wb[tab]
        node_key = tabs[tab]
        # add column header
        column = 2
        for tech in treatment_technologies:
            columnHeaderCellLocation = f'{get_column_letter(column)}{2}'
            ws[columnHeaderCellLocation] = tech
            column+=1
        row = 3
        _print(f'{tab}: adding {node_key}')
        for node in data[node_key]:
            cellLocation = f'{get_column_letter(1)}{row}'
            ws[cellLocation] = node
            row+=1
    
    tabs = {
        "TreatmentOperationalCost": "TreatmentSites",
        "TreatmentEfficiency": "TreatmentSites",
        "RemovalEfficiency": "TreatmentSites",
    }
    column = 2
    for tab in tabs:
        ws = wb[tab]
        node_key = tabs[tab]
        row = 3
        for technology in treatment_technologies:
            value = treatment_technologies[technology][tab]
            _print(f'{tab}: adding {technology}')
            for node in data[node_key]:
                treatmentCellLocation = f'{get_column_letter(column-1)}{row}'
                technologyCellLocation = f'{get_column_letter(column)}{row}'
                valueCellLocation = f'{get_column_letter(column+1)}{row}'
                ws[treatmentCellLocation] = node
                ws[technologyCellLocation] = technology
                ws[valueCellLocation] = value
                row+=1

    tabs = {
        "TreatmentExpansionLeadTime": "TreatmentSites"
    }
    
    for tab in tabs:
        ws = wb[tab]
        node_key = tabs[tab]
        row = 3
        column = 3
        # add column header
        for cap in treatment_capacities:
            columnHeaderCellLocation = f'{get_column_letter(column)}{2}'
            ws[columnHeaderCellLocation] = cap
            column+=1
        for technology in treatment_technologies:
            _print(f'{tab}: adding {technology}')
            for node in data[node_key]:
                treatmentCellLocation = f'{get_column_letter(1)}{row}'
                technologyCellLocation = f'{get_column_letter(2)}{row}'
                ws[treatmentCellLocation] = node
                ws[technologyCellLocation] = technology
                i = 3
                for treatmentCapacity in treatment_capacities:
                    value = treatment_capacities[treatmentCapacity][tab]
                    cellLocation = f'{get_column_letter(i)}{row}'
                    ws[cellLocation] = value
                    i+=1
                row+=1

    tabs = {
        "TreatmentExpansionCost": {"node": "TreatmentSites", "default": "TreatmentTechnologies"}
    }
    
    for tab in tabs:
        ws = wb[tab]
        node_key = tabs[tab]["node"]
        row = 3
        column = 3
        # add column headers
        for cap in treatment_capacities:
            columnHeaderCellLocation = f'{get_column_letter(column)}{2}'
            ws[columnHeaderCellLocation] = cap
            column+=1
        
        for default in defaults[tabs[tab]["default"]]:
            default_values = defaults[tabs[tab]["default"]][default]
            _print(f'{tab}: adding {default}')
            for node in data[node_key]:
                treatmentCellLocation = f'{get_column_letter(1)}{row}'
                technologyCellLocation = f'{get_column_letter(2)}{row}'
                ws[treatmentCellLocation] = node
                ws[technologyCellLocation] = default
                i = 3
                value = default_values[tab]
                for treatmentCapacity in treatment_capacities:
                    valueCellLocation = f'{get_column_letter(i)}{row}'
                    ws[valueCellLocation] = value
                    i+=1
                row+=1


    tabs = {
        "PipelineCapexCapacityBased": {"node": "connections", "default": "PipelineDiameters"},
        "PipelineExpansionLeadTime_Capac": {"node": "connections", "default": "PipelineDiameters"},
    }
    for tab in tabs:
        ws = wb[tab]
        node_key = tabs[tab]["node"]
        row = 2
        column = 3
        default_values = defaults[tabs[tab]["default"]]
        # add column headers
        for each in default_values:
            cellLocation = f'{get_column_letter(column)}{row}'
            ws[cellLocation] = each
            column+=1
        row = 3
        for location in data[node_key]['all_connections']:
            for destination in data[node_key]['all_connections'][location]:
                _print(f'{tab}: adding {location}:{destination} arc')
                locationCellLocation = f'{get_column_letter(1)}{row}'
                destinationCellLocation = f'{get_column_letter(2)}{row}'
                ws[locationCellLocation] = location
                ws[destinationCellLocation] = destination
                row+=1

    ## final step: Save and close
    wb.save(excel_path)
    wb.close()
    return "success"

def determineConnectionsFromArcs(data):
    ## This is a work in progress
    arcs = data.get("arcs", {})
    connections = {
        "all_connections": {}
    }
    data["connections"] = connections
    for arc_key in arcs:
        arc = arcs[arc_key]
        nodes = arc.get("nodes")
        for connecting_node in nodes:
            connecting_node_name = connecting_node["name"]
            outgoing_nodes = connecting_node.get("outgoing_nodes", [])
            connections["all_connections"][connecting_node_name] = outgoing_nodes


    return data

def PreprocessMapData(map_data):
    all_nodes = map_data.get("all_nodes", {})
    default_node_type = map_data.get("defaultNode", "NetworkNode") ## default to network node if we don't have a default node type

    excel_data = {
        'all_nodes': all_nodes,
        'arcs': map_data["arcs"],
        'polygons': map_data["polygons"],
        'ProductionPads': {},
        'CompletionsPads': {},
        'NetworkNodes': {},
        'SWDSites': {},
        'TreatmentSites': {},
        'StorageSites': {},
        'FreshwaterSources': {},
        'ReuseOptions': {},
        'connections': {}
    }

    node_types = {
        'ProductionPad': 'ProductionPads',
        'CompletionsPad': 'CompletionsPads',
        'NetworkNode': 'NetworkNodes',
        'DisposalSite': 'SWDSites',
        'TreatmentSite': 'TreatmentSites',
        'StorageSite': 'StorageSites',
        'FreshwaterSource': 'FreshwaterSources',
        'ReuseOption': 'ReuseOptions'
    }

    for node in all_nodes:
        node_data = all_nodes[node]
        node_type = node_data.get("node_type", default_node_type)
        excel_key = node_types.get(node_type, None)
        if excel_key is None:
            _print(f"found faulty node type: {node_type}")
            continue
        excel_data[excel_key][node] = node_data
    
    _print(f"finished adding nodes to excel data")

    data = determineConnectionsFromArcs(excel_data)

    return data



# sample_arcs = {
#     "p2": {
#         "styleUrl": "#m_ylw-pushpin",
#         "tessellate": "1",
#         "coordinates": [
#             ["-103.8539378936199", "34.53213547776518", "0"],
#             ["-103.8473422508395", "34.49228306568203", "0"],
#             ["-103.8223048221497", "34.49098515301236", "0"],
#         ],
#         "LineString": "",
#         "node_type": "path",
#         "node_list": ["N01", "K01", "K01"],
#         "nodes": [
#             {
#                 "name": "N01",
#                 "outgoing_nodes": ["K01"],
#                 "coordinates": ["-103.8539378936199", "34.53213547776518", "0"],
#             },
#             {
#                 "name": "K01",
#                 "outgoing_nodes": ["K02", "N01"],
#                 "coordinates": ["-103.8473422508395", "34.49228306568203", "0"],
#             },
#             {
#                 "name": "K02",
#                 "outgoing_nodes": ["K01"],
#                 "coordinates": ["-103.8223048221497", "34.49098515301236", "0"],
#             },
#         ],
#     },

# }

# data_with_connections = determineConnectionsFromArcs({"arcs": sample_arcs})
# _print(data_with_connections)