import logging
import shutil
import os
import datetime
import math
import time
from pathlib import Path
import tinydb
from fastapi import HTTPException
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from importlib.resources import files

import idaes.logger as idaeslog
from pareto.utilities.get_data import get_display_units

from app.internal.get_data import get_data, get_input_lists
from app.internal.settings import AppSettings

# _log = idaeslog.getLogger(__name__)
_log = logging.getLogger(__name__)

class ScenarioHandler:
    """Manage the saved scenarios."""

    VERSION = 2
    SCENARIO_DB_FILE = f"scenarios.json"
    LOCKED = False

    def __init__(self, **kwargs) -> None:

        _log.info(f"initializing scenario handler")
        self.app_settings = AppSettings(**kwargs)

        self.scenario_list = {}
        self.next_id = 0
        self.background_tasks = []
        self.data_directory_path = self.app_settings.data_basedir
        self.scenarios_path = self.data_directory_path / f"v{self.VERSION}" / self.SCENARIO_DB_FILE
        self.excelsheets_path = self.data_directory_path / f"v{self.VERSION}" / "excelsheets"
        self.outputs_path = self.data_directory_path / f"v{self.VERSION}" / "outputs"
        self.input_diagrams_path = self.data_directory_path / f"v{self.VERSION}" / "input_diagrams"
        self.output_diagrams_path = self.data_directory_path / f"v{self.VERSION}" / "output_diagrams"

        _log.info(f"app directory: {os.path.dirname(os.path.abspath(__file__))}")
        _log.info(f"currently operating in directory: {os.getcwd()}")
        try:
            _log.info(f"changing to home directroy")
            home_dir = Path.home()
            _log.info(f"new directory: {home_dir}")
            os.chdir(home_dir)
        except Exception as e:
            _log.info(f"unable to change to app directroy: {e}")

        ### check for data directory
        has_data = os.path.isdir(self.data_directory_path / f"v{self.VERSION}")
        _log.info(f'has_data is {has_data}')


        # create data directories
        _log.info(f"making directory: {self.data_directory_path}")
        self.data_directory_path.mkdir(parents=True, exist_ok=True)

        _log.info(f"making directory: {self.excelsheets_path}")
        self.excelsheets_path.mkdir(parents=True, exist_ok=True)

        _log.info(f"making directory: {self.outputs_path}")
        self.outputs_path.mkdir(parents=True, exist_ok=True)

        _log.info(f"making directory: {self.input_diagrams_path}")
        self.input_diagrams_path.mkdir(parents=True, exist_ok=True)

        _log.info(f"making directory: {self.output_diagrams_path}")
        self.output_diagrams_path.mkdir(parents=True, exist_ok=True)

        # if not has_data:
        #     _log.info('importing default data')
        #     try:
        #         self.import_default_data()
        #     except Exception as e:
        #         _log.error(f'unable to import default data: {e}')

        # Connect to DB
        path = self.scenarios_path
        self._db = tinydb.TinyDB(path)
        self.update_next_id()
        self.retrieve_scenarios()

    def import_default_data(self):
        default_ids = [1,2,3,4,5]
        default_directories = [("excelsheets", "xlsx"), ("input_diagrams", "png"), ("output_diagrams", "png"), ("outputs", "xlsx")]
        default_data_path = files('app').joinpath(f"internal/assets/v{self.VERSION}_default")
        destination_directory = self.data_directory_path / f"v{self.VERSION}"
        for directory in default_directories:
            d_name = directory[0]
            d_ext  = directory[1]
            for each in default_ids:
                f_path = default_data_path.joinpath(f"{d_name}/{each}.{d_ext}")
                destination_path = destination_directory.joinpath(f"{d_name}/{each}.{d_ext}")
                _log.info(f'moving {f_path} to {destination_path}')
                shutil.copyfile(f_path, destination_path)
        ### copy over scenarios.json files
        f_path = default_data_path.joinpath(f"scenarios.json")
        destination_path = destination_directory.joinpath(f"scenarios.json")
        _log.info(f'moving {f_path} to {destination_path}')
        shutil.copyfile(f_path, destination_path)

    def retrieve_scenarios(self):
        _log.info(f"retrieving scenarios")
        # check if db is in use. if so, wait til its done being used
        locked = self.LOCKED
        while(locked):
            time.sleep(0.5)
            locked = self.LOCKED
        self.LOCKED = True
        query = tinydb.Query()
        scenarios = self._db.search((query.id_ != None) & (query.version == self.VERSION))
        scenario_list = {}
        _log.info(f'found {len(scenarios)} scenarios')
        if len(scenarios) > 0:
            for each in scenarios:
                scenario_list[each["id_"]] =  each['scenario']
            self.scenario_list=scenario_list
        else:
            self.scenario_list={}


        self.LOCKED = False
        
    def update_scenario(self, updatedScenario):
        _log.info(f"Updating scenario list")

        try:
            id_ = updatedScenario["id"]
            newName = updatedScenario["name"]
            oldName = self.scenario_list[id_]["name"]
            # _log.info('trying to replace input diagram for scenarios renamed to SRA')
            # _log.info(f'new scenario name is: {newName}')
            # _log.info(f'previous name was {oldName}')
            if newName.upper() == "SRA" and oldName.upper() != "SRA":
                ## move SRA image into input network diagram spot
                _log.info('scenario has been RENAMED to SRA')
                input_diagramFileType = 'png'
                original_input_diagram_path = f'{os.path.dirname(os.path.abspath(__file__))}/assets/SRAInput.png'
                new_input_diagram_path = f"{self.input_diagrams_path}/{id_}.{input_diagramFileType}"
                shutil.copyfile(original_input_diagram_path, new_input_diagram_path)
                updatedScenario[f'inputDiagramExtension'] = 'png'
            if newName.upper() == "WORKSHOP BASELINE" and oldName.upper() != "WORKSHOP BASELINE":
                input_diagramFileType = 'png'
                original_input_diagram_path = f'{os.path.dirname(os.path.abspath(__file__))}/assets/workshop_baseline_input.png'
                new_input_diagram_path = f"{self.input_diagrams_path}/{id_}.{input_diagramFileType}"
                shutil.copyfile(original_input_diagram_path, new_input_diagram_path)
                updatedScenario[f'inputDiagramExtension'] = 'png'
            if newName.upper() == "WORKSHOP SRA" and oldName.upper() != "WORKSHOP SRA":
                input_diagramFileType = 'png'
                original_input_diagram_path = f'{os.path.dirname(os.path.abspath(__file__))}/assets/workshop_SRA_input.png'
                new_input_diagram_path = f"{self.input_diagrams_path}/{id_}.{input_diagramFileType}"
                shutil.copyfile(original_input_diagram_path, new_input_diagram_path)
                updatedScenario[f'inputDiagramExtension'] = 'png'
            if newName.upper() == "WORKSHOP BENEFICIAL REUSE" and oldName.upper() != "WORKSHOP BENEFICIAL REUSE":
                input_diagramFileType = 'png'
                original_input_diagram_path = f'{os.path.dirname(os.path.abspath(__file__))}/assets/workshop_beneficial_reuse_input.png'
                new_input_diagram_path = f"{self.input_diagrams_path}/{id_}.{input_diagramFileType}"
                shutil.copyfile(original_input_diagram_path, new_input_diagram_path)
                updatedScenario[f'inputDiagramExtension'] = 'png'
        except:
            _log.error('unable to check and/or replace input diagram for SRA')

        # check if db is in use. if so, wait til its done being used
        locked = self.LOCKED
        while(locked):
            time.sleep(0.5)
            locked = self.LOCKED
        self.LOCKED = True
        query = tinydb.Query()
        self._db.upsert(
            {"scenario": updatedScenario, 'id_': updatedScenario['id'], 'version': self.VERSION},
            ((query.id_ == updatedScenario['id']) & (query.version == self.VERSION)),
        )

        self.LOCKED = False

        self.retrieve_scenarios()

        return updatedScenario
    
    def upload_excelsheet(self, output_path, scenarioName, filename, kmz_data=None):
        _log.info(f"Uploading excel sheet: {scenarioName}")

        [set_list, parameter_list] = get_input_lists()

        # read in data from uploaded excel sheet
        [df_sets, df_parameters, frontend_parameters] = get_data(output_path, set_list, parameter_list)
        
        try:
            display_units = get_display_units(parameter_list, df_parameters["Units"])
        except Exception as e:
            _log.error(f'unable to get units: {e}')
            display_units = {}

        del frontend_parameters['Units']

        # convert tuple keys into dictionary values - necessary for javascript interpretation
        for key in df_parameters:
            original_value = df_parameters[key]
            new_value=[]
            for k, v in original_value.items():
                try:
                    if math.isnan(v):
                        new_value.append({'key':k, 'value': ''})
                    else:
                        new_value.append({'key':k, 'value': v})
                except:
                    new_value.append({'key':k, 'value': v})
            df_parameters[key] = new_value

        # convert pandas series into lists
        for key in df_sets:
            df_sets[key] = df_sets[key].tolist()

        # create scenario object
        current_day = datetime.date.today()
        date = datetime.date.strftime(current_day, "%m/%d/%Y")
        return_object = {
            "name": scenarioName, 
            "id": self.next_id, 
            "date": date,
            "data_input": {"df_sets": df_sets, "df_parameters": frontend_parameters, 'display_units': display_units, "map_data": kmz_data}, 
            "optimization": 
                {
                    "objective":"cost", 
                    "runtime": 900, 
                    "pipelineCostCalculation": "distance_based", 
                    "waterQuality": "false", 
                    "hydraulics": "false",
                    "solver": "cbc",
                    "build_units": "scaled_units",
                    "optimalityGap": 0,
                    "scale_model": True
                }, 
            "results": {"status": "Draft", "data": {}},
            "override_values": 
                {
                    "vb_y_overview_dict": {},
                    "v_F_Piped_dict": {},
                    "v_F_Sourced_dict": {},
                    "v_F_Trucked_dict": {},
                    "v_L_Storage_dict": {},
                    "v_L_PadStorage_dict": {},
                    "vb_y_Pipeline_dict": {},
                    "vb_y_Disposal_dict": {},
                    "vb_y_Storage_dict": {},
                    "vb_y_Treatment_dict": {}
                }
            }

        # check if db is in use. if so, wait til its done being used
        locked = self.LOCKED
        while(locked):
            time.sleep(0.5)
            locked = self.LOCKED
        self.LOCKED = True
        self._db.insert({'id_': self.next_id, "scenario": return_object, 'version': self.VERSION})
        self.LOCKED = False
        
        return_object = self.check_for_diagram(self.next_id, filename.split('.')[0])

        self.update_next_id()
        self.retrieve_scenarios()
        
        return return_object
    
    def check_for_diagram(self, id, filename = None):
        scenario = self.get_scenario(id)
        extension = "png"
        if filename:
            diagramType = "input"
            diagramIdentifier = filename
            scenario_diagram_path = f"{self.input_diagrams_path}/{id}.{extension}"
        else:
            diagramType = "output"
            diagramIdentifier = scenario['name']
            scenario_diagram_path = f"{self.output_diagrams_path}/{id}.{extension}"
        try:
            if scenario['name'].upper() == "WORKSHOP BASELINE":
                diagram_path = f'{os.path.dirname(os.path.abspath(__file__))}/assets/workshop_baseline_{diagramType}.png'
            elif scenario['name'].upper() == "WORKSHOP SRA":
                diagram_path = f'{os.path.dirname(os.path.abspath(__file__))}/assets/workshop_sra_{diagramType}.png'
            elif scenario['name'].upper() == "WORKSHOP BENEFICIAL REUSE":
                diagram_path = f'{os.path.dirname(os.path.abspath(__file__))}/assets/workshop_beneficial_reuse_{diagramType}.png'
            else: 
                diagram_path = files('pareto').parent.joinpath(f"docs/img/{diagramIdentifier}.{extension}")
            _log.info(f'diagram_path is : {diagram_path}')
            shutil.copyfile(diagram_path, scenario_diagram_path)
            scenario[f"{diagramType}DiagramExtension"] = extension
            return self.update_scenario(scenario)
        except Exception as e:
            _log.info(f'unable to find diagram_path: {e}')
            return scenario


    def copy_scenario(self, id, new_scenario_name):
        _log.info(f"copying scenario with id: {id}")

        try:
            # copy scenario with given id
            new_scenario = self.scenario_list[id].copy()
            new_scenario_id = self.next_id

            # update scenario name, id, and creation date
            current_day = datetime.date.today()
            date = datetime.date.strftime(current_day, "%m/%d/%Y")
            new_scenario["name"] = new_scenario_name
            new_scenario["id"] = new_scenario_id
            new_scenario["date"] = date

            # create copy of excel sheet input
            original_excel_path = "{}/{}.xlsx".format(self.excelsheets_path,id)
            new_excel_path = "{}/{}.xlsx".format(self.excelsheets_path,new_scenario_id)
            shutil.copyfile(original_excel_path, new_excel_path)

            # create copy of excel sheet output (if it exists)
            original_output_path = "{}/{}.xlsx".format(self.outputs_path,id)
            new_output_path = "{}/{}.xlsx".format(self.outputs_path,new_scenario_id)
            if (os.path.isfile(original_output_path)):
                shutil.copyfile(original_output_path, new_output_path)

            # check for disposal override scenario. if found, insert disposal diagrams in. otherwise, copy input diagram over
            if new_scenario_name.lower() == 'disposal override':
                output_diagramFileType = 'png'
                original_output_diagram_path = f'{os.path.dirname(os.path.abspath(__file__))}/assets/DisposalOverride.png'
                new_output_diagram_path = f"{self.output_diagrams_path}/{new_scenario_id}.{output_diagramFileType}"

                input_diagramFileType = 'png'
                original_input_diagram_path = f'{os.path.dirname(os.path.abspath(__file__))}/assets/DisposalOverrideInput.png'
                new_input_diagram_path = f"{self.input_diagrams_path}/{new_scenario_id}.{input_diagramFileType}"
                try:
                    shutil.copyfile(original_output_diagram_path, new_output_diagram_path)
                    new_scenario[f'outputDiagramExtension'] = 'png'
                except Exception as e:
                    _log.error(f'unable to copy disposal override output diagram over: {e}')
                try:
                    shutil.copyfile(original_input_diagram_path, new_input_diagram_path)
                    new_scenario[f'inputDiagramExtension'] = 'png'
                except Exception as e:
                    _log.error(f'unable to copy disposal override input diagram over: {e}')
            else:
                try:
                    input_diagramFileType = self.scenario_list[id][f'inputDiagramExtension']
                    original_input_diagram_path = f"{self.input_diagrams_path}/{id}.{input_diagramFileType}"
                    new_input_diagram_path = f"{self.input_diagrams_path}/{new_scenario_id}.{input_diagramFileType}"
                    if (os.path.isfile(original_input_diagram_path)):
                        shutil.copyfile(original_input_diagram_path, new_input_diagram_path)
                except:
                    _log.info('unable to copy input network diagram')
            


            # add record in db for new scenario
            # check if db is in use. if so, wait til its done being used
            locked = self.LOCKED
            while(locked):
                time.sleep(0.5)
                locked = self.LOCKED
            self.LOCKED = True
            self._db.insert({'id_': new_scenario_id, "scenario": new_scenario, 'version': self.VERSION})
            self.LOCKED = False
            self.update_next_id()
            self.retrieve_scenarios()

            # return updated scenario list
            return self.scenario_list, new_scenario_id

        except Exception as e:
            _log.error(f"error copying scenario: {e}")
            raise HTTPException(
                    500, f"unable to make copy of scenario with id {id}: {e}"
                )

    def delete_scenario(self, index):
        _log.info(f"Deleting scenario #{index}")
        # check if db is in use. if so, wait til its done being used
        locked = self.LOCKED
        while(locked):
            time.sleep(0.5)
            locked = self.LOCKED
        self.LOCKED = True
        try:
            index = int(index)
            query = tinydb.Query()
            self._db.remove((query.id_ == index) & (query.version == self.VERSION))
            self.LOCKED = False

            # remove input excel sheet
            excel_sheet = "{}/{}.xlsx".format(self.excelsheets_path,index)
            os.remove(excel_sheet)
        except Exception as e:
            _log.error(f"unable to delete scenario #{index}: {e}")

        # remove output excel sheet
        try:
            excel_sheet = "{}/{}.xlsx".format(self.outputs_path,index)
            os.remove(excel_sheet)
        except Exception as e:
            _log.error(f"unable to remove output for #{index}: {e}")

        # remove diagram images (if they exist)
        try:
            diagramFileType = self.scenario_list[index]['inputDiagramExtension']
            input_diagram = f"{self.input_diagrams_path}/{index}.{diagramFileType}"
            _log.info(f'removing input diagram from location: {input_diagram}')
            if os.path.isfile(input_diagram):
                os.remove(input_diagram)
        except Exception as e:
            _log.error(f"unable to remove input diagram for #{index}: {e}")
        try:
            diagramFileType = self.scenario_list[index]['outputDiagramExtension']
            output_diagram = f"{self.output_diagrams_path}/{index}.{diagramFileType}"
            _log.info(f'removing output diagram from location: {output_diagram}')
            if os.path.isfile(output_diagram):
                os.remove(output_diagram)
        except Exception as e:
            _log.error(f"unable to remove output diagram for #{index}: {e}")

        # # remove completions demand plot
        # try:
        #     plot_file = "{}/{}_plot.html".format(self.outputs_path,index)
        #     os.remove(plot_file)
        # except Exception as e:
        #     _log.error(f"unable to delete completions demand plot for #{index}: {e}")

        # update scenario list
        self.retrieve_scenarios()

    def get_scenario(self, id):
        try:
            # check if db is in use. if so, wait til its done being used
            locked = self.LOCKED
            while(locked):
                time.sleep(0.5)
                locked = self.LOCKED
            self.LOCKED = True
            query = tinydb.Query()
            scenario = self._db.search((query.id_ == id) & (query.version == self.VERSION))
            self.LOCKED = False
            return scenario[0]['scenario']
        except Exception as e:
            _log.error(f'unable to get scenario with id {id}: {e}')
            return {'error': f'unable to get scenario with id {id}: {e}'}

    def get_plots(self, id):
        return_object = {}
        plot_files = {
            'CompletionsDemand': 'Completion Pad Demand', 
            "PadRates": "Production Forecast", 
            "FlowbackRates": "Flowback Forecast"
        }
        for each in plot_files:
            try:
                key = plot_files[each]
                plot_file = f"{self.outputs_path}/{id}_{each}_plot.html"
                with open(plot_file, 'r') as f:
                    plot_html = f.read()
                    return_object[key] = plot_html
            except Exception as e:
                _log.error(f"unable to get plot for id{id}: {e}")
                raise HTTPException(
                    500, f"unable to find plot: {e}"
                )
        return return_object

    def add_background_task(self, id):
        self.background_tasks.append(id)
        return self.background_tasks

    def remove_background_task(self, id):
        if id in self.background_tasks: 
            self.background_tasks.remove(id)
        else:
            _log.error(f'id #{id} is not in background tasks list')
        return self.background_tasks

    def get_list(self):
        return self.scenario_list

    def get_next_id(self):
        nextid = self.next_id
        return nextid

    def get_excelsheet_path(self, id):
        return f"{self.excelsheets_path}/{id}.xlsx"

    def get_diagram(self, diagram_type, id):
        try:
            diagramFileType = self.scenario_list[id][f'{diagram_type}DiagramExtension']
            if diagram_type == "input":
                diagramLocation = f"{self.input_diagrams_path}/{id}.{diagramFileType}"
            elif diagram_type == "output":
                diagramLocation = f"{self.output_diagrams_path}/{id}.{diagramFileType}"
            if os.path.isfile(diagramLocation):
                return diagramLocation
            else:
                _log.error(f"unable to find diagram for id {id}")
                raise HTTPException(400, detail=f"no diagram found")
        except Exception as e:
            _log.error(f"error: unable to find diagram for id {id}: {e}")
            raise HTTPException(400, detail=f"no diagram found: {e}")
        
    def update_next_id(self):
        try:
            # check if db is in use. if so, wait til its done being used
            locked = self.LOCKED
            while(locked):
                time.sleep(0.5)
                locked = self.LOCKED
            self.LOCKED = True

            query = tinydb.Query()
            # el = self._db.search((query.version == self.VERSION))[-1]
            # next_id = el.doc_id+1
            db_items = self._db.search((query.version == self.VERSION))
            highest_id = 0
            for each in db_items:
                if each["id_"] > highest_id:
                    highest_id = each["id_"]
            next_id = highest_id + 1
            _log.info(f'setting next id: {next_id}')
            self.next_id = next_id
        except Exception as e:
            _log.info(f"no documents found; next id is 0")
            _log.error(f"{e}")

        self.LOCKED = False

    def get_background_tasks(self):
        return self.background_tasks

    def update_excel(self, id, table_key, updatedTable):
        _log.info(f'updating id {id} table {table_key}')
        excel_path = self.get_excelsheet_path(id)
        wb = load_workbook(excel_path, data_only=True)
        _log.info(f'loaded workbook from path: {excel_path}')

        x = 1
        y = 3
        ws = wb[table_key]
        # for column in scenario["data_input"]["df_parameters"][table_key]:

        # update excel sheet
        _log.info(f'updating table: {updatedTable}')
        for column in updatedTable:
            y = 3
            for cellValue in updatedTable[column]:
                cellLocation = f'{get_column_letter(x)}{y}'
                originalValue = ws[cellLocation].value
                if cellValue == "":
                    newValue = None
                else:
                    newValue = cellValue
                    # try:
                    #     newValue = int(cellValue)
                    # except ValueError:
                    #     try:
                    #         newValue = float(cellValue)
                    #     except: 
                    #         newValue = cellValue
                if originalValue != newValue:
                    # print('updating value')
                    ws[cellLocation] = newValue
                y+=1
            x+=1
        wb.save(excel_path)
        wb.close()
        _log.info(f'saved workbook')
        # fetch scenario
        try:
            # check if db is in use. if so, wait til its done being used
            locked = self.LOCKED
            while(locked):
                time.sleep(0.5)
                locked = self.LOCKED
            self.LOCKED = True
            query = tinydb.Query()
            scenario = self._db.search((query.id_ == int(id)) & (query.version == self.VERSION))[0]['scenario']
            scenario["data_input"]["df_parameters"][table_key] = updatedTable
        except Exception as e:
            _log.info(f"unable to fetch scenario: {e}")
        self.LOCKED = False
        # update scenario
        return self.update_scenario(scenario)

    def upload_diagram(self, output_path, id, diagram_type):
        # check if db is in use. if so, wait til its done being used
        locked = self.LOCKED
        while(locked):
            time.sleep(0.5)
            locked = self.LOCKED
        self.LOCKED = True
        query = tinydb.Query()
        scenario = self._db.search((query.id_ == int(id)) & (query.version == self.VERSION))[0]['scenario']
        scenario[f"{diagram_type}DiagramExtension"] = output_path.split('.')[-1]
        self._db.upsert(
            {"scenario": scenario, 'id_': scenario['id'], 'version': self.VERSION},
            ((query.id_ == scenario['id']) & (query.version == self.VERSION)),
        )
        self.LOCKED = False
        self.retrieve_scenarios()
        
        return
    
    def delete_diagram(self, diagram_type, index):
        try:
            diagramFileType = self.scenario_list[index][f'{diagram_type}DiagramExtension']
            if diagram_type == "input":
                diagram = f"{self.input_diagrams_path}/{index}.{diagramFileType}"
            elif diagram_type == "output":
                diagram = f"{self.output_diagrams_path}/{index}.{diagramFileType}"
            _log.info(f'removing diagram from location: {diagram}')
            if os.path.isfile(diagram):
                os.remove(diagram)
        except Exception as e:
            _log.error(f"unable to remove diagram for #{index}: {e}")
            raise HTTPException(400, detail=f"unable to remove diagram: {e}")
        return self.scenario_list[index]

scenario_handler = ScenarioHandler()
